import folium
import requests
import json
from tkinter import *
import tkinter.ttk as ttk
import tkinter.font as font
import tkinter.filedialog as filedialog
import tkinter.messagebox
import openpyxl
import shapefile
import csv


def getLatLng(addr, crs):
    key = '개발자 키'
    url = 'http://api.vworld.kr/req/address?service=address&request=getCoord&type=road' \
          '&crs=' + crs + '&address=' + addr + '&key=' + key
    try:
        result = json.loads(str(requests.get(url).text))['response']['result']
        a, b = result.get('point').values()
        return float(b), float(a)
    except LookupError:
        url = 'http://api.vworld.kr/req/address?service=address&request=getCoord&type=parcel' \
              '&crs=' + crs + '&address=' + addr + '&key=' + key
        result = json.loads(str(requests.get(url).text))['response']['result']
        a, b = result.get('point').values()
        return float(b), float(a)


def getAddr(lng, lat, crs):
    key = '개발자 키'
    url = 'http://api.vworld.kr/req/address?service=address&request=getAddress&type=both' \
          '&crs=' + crs + '&point=' + lat + ',' + lng + '&key=' + key
    try:
        result = json.loads(str(requests.get(url).text))['response']['result']
        address, road_address = result[0]['text'], result[1]['text']
        return road_address, address
    except LookupError:
        return '0', '0'


def csv_converter(root, filename):
    wb = openpyxl.load_workbook(root + '\\' + filename)
    sh = wb.get_sheet_by_name('Sheet')
    with open(root + '/' + 'buchuck.csv', 'w', newline="") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])


class Map:
    def __init__(self, win):
        self.win = win
        self.win.config(background='lavender')
        self.font = font.Font(family='consolas', size=15, weight='bold')
        self.fontSmaller = font.Font(family='consolas', size=12, weight='bold')
        self.callButton = Button(self.win, text='입력파일', font=self.font, command=self.askRoot)
        self.callButton.place(x=10, y=10, width=150, height=40)
        self.value = ['주소->좌표', '좌표->주소']
        self.changeBar = ttk.Combobox(self.win, values=self.value, font=self.fontSmaller)
        self.changeBar.place(x=200, y=5, width=200, height=25)
        self.changeBar.set(self.value[0])
        self.crs = ['EPSG:4326', 'EPSG:3857', 'EPSG:5174', 'EPSG:5179']
        self.crsBar = ttk.Combobox(self.win, values=self.crs, font=self.fontSmaller)
        self.crsBar.place(x=200, y=40, width=200, height=25)
        self.crsBar.set(self.crs[0])
        self.startButton = Button(self.win, text='변환시작', font=self.font, command=self.run)
        self.startButton.place(x=420, y=10, width=150, height=40)
        self.label = Label(self.win, text='진행률', font=self.font, background='lavender')
        self.label.place(x=580, y=15)
        self.progress = ttk.Progressbar(self.win, maximum=100, mode='determinate')
        self.progress.place(x=650, y=20, width=130, height=20)
        self.label2 = Label(self.win, text='저장 경로', font=self.fontSmaller, background='lavender')
        self.label2.place(x=10, y=70)
        self.root = Text(self.win)
        self.root.place(x=100, y=75, width=600, height=20)
        self.rootQuery = Button(self.win, text='...', command=self.askSaveRoot)
        self.rootQuery.place(x=720, y=75, width=30, height=20)
        self.column = ['zero', 'one', 'two', 'three', 'four', 'five']
        self.chart = ttk.Treeview(self.win, columns=self.column)
        self.chart.place(x=10, y=120, width=780, height=250)
        self.chart.column("#0", width=50)
        self.chart.heading("#0", text="No", anchor="center")
        self.chart.column("#1", width=100)
        self.chart.heading("#1", text="OWNER", anchor="center")
        self.chart.column('#2', width=250)
        self.chart.heading('#2', text='ADDR', anchor='center')
        self.chart.column('#3', width=175)
        self.chart.heading('#3', text='경도', anchor='center')
        self.chart.column('#4', width=175)
        self.chart.heading('#4', text='위도', anchor='center')
        self.vsb = Scrollbar(self.chart, orient='vertical', command=self.chart.yview)
        self.vsb.pack(side='right', fill='y')
        self.list = ['lng', 'lat', '지번주소', '도로명주소']
        self.changeBar.bind('<<ComboboxSelected>>', self.refresh)
        self.first = []
        self.second = []
        self.third = []
        self.lat = []
        self.lng = []
        self.roadAddr = []
        self.address = []
        self.map_info = folium.Map(
                    zoom_start=13
                )

    def askRoot(self):
        root = filedialog.askopenfilename(initialdir='C:')
        wb = openpyxl.load_workbook(root)
        sheet = wb.active
        self.first = []
        self.second = []
        self.third = []
        if self.changeBar.get() == self.value[0]:
            for i in range(0, sheet.max_row):
                self.first.append(sheet.cell(row=i + 1, column=1).value)
                self.second.append(sheet.cell(row=i + 1, column=2).value)
                self.inputText(self.first, self.second, None, None, None)
            wb.close()
        else:
            for i in range(0, sheet.max_row):
                self.first.append(sheet.cell(row=i + 1, column=1).value)
                self.second.append(sheet.cell(row=i + 1, column=2).value)
                self.third.append(sheet.cell(row=i + 1, column=3).value)
                self.inputText(self.first, self.second, self.third, None, None)
            wb.close()

    def inputText(self, first, second, third, fourth, fifth):
        self.clearchart()
        i = 0
        if first and second and third is None and fourth is None and fifth is None:
            for j, k in zip(first, second):
                i += 1
                self.chart.insert('', 'end', text=i, values=(j, k))

        if first and second and third and fourth is None and fifth is None:
            for j, k, l in zip(first, second, third):
                i += 1
                self.chart.insert('', 'end', text=i, values=(j, k, l))

        if first and second and third and fourth and fifth is None:
            for j, k, l, m in zip(first, second, third, fourth):
                i += 1
                self.chart.insert('', 'end', text=i, values=(j, k, l, m))

        if first and second and third and fourth and fifth:
            for j, k, l, m, n in zip(first, second, third, fourth, fifth):
                i += 1
                self.chart.insert('', 'end', text=i, values=(j, k, l, m, n))

    def askSaveRoot(self):
        root = filedialog.askdirectory()
        self.root.delete(1.0, END)
        self.root.insert(1.0, root)

    def clearchart(self):
        self.chart.delete(*self.chart.get_children())

    def refresh(self, event):
        if self.changeBar.get() == self.value[0]:
            self.clearchart()
            self.chart.column("#0", width=50)
            self.chart.heading("#0", text="No", anchor="center")
            self.chart.column("#1", width=100)
            self.chart.heading("#1", text="OWNER", anchor="center")
            self.chart.column('#2', width=250)
            self.chart.heading('#2', text='ADDR', anchor='center')
            self.chart.column('#3', width=175)
            self.chart.heading('#3', text='경도', anchor='center')
            self.chart.column('#4', width=175)
            self.chart.heading('#4', text='위도', anchor='center')

        if self.changeBar.get() == self.value[1]:
            self.clearchart()
            self.chart.column("#0", width=50)
            self.chart.heading("#0", text="No", anchor="center")
            self.chart.column("#1", width=100)
            self.chart.heading("#1", text="OWNER", anchor="center")
            self.chart.column('#2', width=100)
            self.chart.heading('#2', text='lng', anchor='center')
            self.chart.column('#3', width=100)
            self.chart.heading('#3', text='lat', anchor='center')
            self.chart.column('#4', width=200)
            self.chart.heading('#4', text='지번주소', anchor='center')
            self.chart.column('#5', width=200)
            self.chart.heading('#5', text='도로명주소', anchor='center')

    def makeshp(self, judge, root):
        if judge == 0:
            city_shp = shapefile.Writer(root + '\\' + './buchuckOne')
            city_shp.field("No", "C")
            city_shp.field("OWNER", 'C')
            city_shp.field('ADDR', 'C')

            with open(root + '/' + 'buchuck.csv') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')

                for i in reader:
                    No = i[0]
                    OWNER = i[1]
                    ADDR = i[2]
                    lat = float(i[3])
                    lng = float(i[4])

                    city_shp.point(float(lng), float(lat))
                    city_shp.record(No, OWNER, ADDR)
        
        else:
            city_shp = shapefile.Writer(root + '\\' + './buchuckTwo')
            city_shp.field("No", "C")
            city_shp.field("OWNER", 'C')
            city_shp.field("addr", 'C')
            city_shp.field('roadAddr', 'C')

            with open(root + '/' + 'buchuck.csv') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')

                for i in reader:
                    No = i[0]
                    OWNER = i[1]
                    lng = float(i[2])
                    lat = float(i[3])
                    addr = i[4]
                    roadAddr = i[5]

                    city_shp.point(float(lng), float(lat))
                    city_shp.record(No, OWNER, addr, roadAddr)

    def run(self):
        if str(self.root.get(1.0, END)).strip():
            save_root = str(self.root.get(1.0, END)).strip()
        else:
            tkinter.messagebox.showinfo('경고', '저장 경로 설정해')
            return 0
        self.lat = []
        self.lng = []
        crs = self.crsBar.get()
        level = 100 / len(self.second)
        count = 0
        wb = openpyxl.Workbook()
        sheet = wb.active

        if self.changeBar.get() == self.value[0]:
            for i, j in zip(self.second, self.first):
                count += 1
                lat, lng = getLatLng(i, crs)
                self.lat.append(float(lat))  # 변환 코드
                self.lng.append(float(lng))
                self.progress['value'] = int(count*level)
                sheet.cell(row=count, column=1).value = count
                sheet.cell(row=count, column=2).value = j
                sheet.cell(row=count, column=3).value = i
                sheet.cell(row=count, column=4).value = lat
                sheet.cell(row=count, column=5).value = lng
                self.win.update()
                folium.Marker(
                    location=[lat, lng],
                    popup=i,
                    icon=folium.Icon(color='red', icon='info-sign')
                ).add_to(self.map_info)

                folium.LayerControl().add_to(self.map_info)

            self.map_info.save(save_root + '\\' + '주소~좌표 변환.html')
            wb.save(save_root + '/' + '주소~좌표.xlsx')
            csv_converter(save_root, '주소~좌표.xlsx')
            self.makeshp(0, save_root)
            self.inputText(self.first, self.second, self.lng, self.lat, None)

        else:
            for i, j, k in zip(self.second, self.third, self.first):
                count += 1
                roadAddr, address = getAddr(str(j), str(i), crs)
                self.roadAddr.append(roadAddr)
                self.address.append(address)
                self.progress['value'] = int(count*level)
                self.win.update()
                sheet.cell(row=count, column=1).value = count
                sheet.cell(row=count, column=2).value = k
                sheet.cell(row=count, column=3).value = i
                sheet.cell(row=count, column=4).value = j
                sheet.cell(row=count, column=5).value = roadAddr
                sheet.cell(row=count, column=6).value = address

                folium.Marker(
                    location=[j, i],
                    popup=roadAddr,
                    icon=folium.Icon(color='red', icon='info-sign')
                ).add_to(self.map_info)

                folium.LayerControl().add_to(self.map_info)

            self.map_info.save(save_root + '\\' + '좌표~주소 변환.html')
            wb.save(save_root + '/' + '좌표~주소.xlsx')
            csv_converter(save_root, '좌표~주소.xlsx')
            self.makeshp(1, save_root)
            self.inputText(self.first, self.second, self.third, self.address, self.roadAddr)


root = Tk()
root.geometry('800x400+100+100')
root.title('붐척이')
root.resizable(False, False)

if __name__ == '__main__':
    Map(root)
    root.mainloop()
